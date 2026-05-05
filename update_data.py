#!/usr/bin/env python3
"""
HITACHI Rail T&C Portal - Data Update Script

Run this whenever your CSV files OR TestPlan_Master.xlsm change.
It reads everything and regenerates data.js.

USAGE:
    1. Place latest files in the same folder as this script:
       - CX_DATA--_Line_Item.csv
       - CX_DATA--_Action_Plan.csv
       - CX_DATA--_Punch_List.csv
       - TestPlan_Master.xlsm
    2. Run: python3 update_data.py
    3. Upload the new data.js to GitHub

Status mappings applied automatically:
    - Line Item Status: Closed -> Passed, Delayed -> Failed
    - Action Plan Status: Pending Test Report Acceptance -> Closed
"""

import json
import os
import sys
from datetime import datetime

# === CONFIGURATION ===
CSV_LINE_ITEM = "CX_DATA--_Line_Item.csv"
CSV_ACTION_PLAN = "CX_DATA--_Action_Plan.csv"
CSV_PUNCH_LIST = "CX_DATA--_Punch_List.csv"
XLSM_TESTPLAN = "TestPlan_Master.xlsm"

OUTPUT_FILE = "data.js"

# === ORG CHART (edit when team changes) ===
ORG = [
    {"name": "Christopher Burford", "title": "Testing & Commissioning Manager", "id": "tc-mgr", "reports": None, "level": 0},
    {"name": "Alex Khoury / Syed Rahman", "title": "Deputy T&C Manager", "id": "deputy", "reports": "tc-mgr", "level": 1},
    {"name": "John Sterrett", "title": "Lead ATS T&C Engineer", "id": "ats", "reports": "deputy", "level": 2},
    {"name": "Thomas Giraud", "title": "Lead CBTC T&C Engineer", "id": "cbtc-lead", "reports": "deputy", "level": 2},
    {"name": "Trevor Abeldt", "title": "CBTC T&C Engineer", "id": "cbtc-eng", "reports": "cbtc-lead", "level": 3},
    {"name": "Srikanth Manickam", "title": "Lead Comms T&C Engineer", "id": "comms-lead", "reports": "deputy", "level": 2},
    {"name": "Roberto Mercurio", "title": "Sr. Comms T&C Engineer", "id": "comms-sr", "reports": "comms-lead", "level": 3},
    {"name": "Viktor Hryshko", "title": "Lead IXL T&C Engineer", "id": "ixl", "reports": "deputy", "level": 2},
    {"name": "Adam Piotrowski / Jimmy Young / Yaroslav Pryimak", "title": "Sr. Signaling T&C Engineer", "id": "sig-sr", "reports": "deputy", "level": 2},
    {"name": "Oleksii Syplyvyi", "title": "Signaling T&C Engineer", "id": "sig-eng", "reports": "sig-sr", "level": 3},
    {"name": "Giancarlo Gaudioso", "title": "Lead Security & Control T&C Engineer", "id": "sec", "reports": "deputy", "level": 2},
    {"name": "TBD", "title": "Lead System T&C Engineer", "id": "sys", "reports": "deputy", "level": 2},
    {"name": "TBD", "title": "Lead Testing Operator", "id": "ops", "reports": "deputy", "level": 2},
]

# === FIELD USERS (edit when team changes) ===
# Names + 4-digit PINs for field team to access logging form.
FIELD_USERS = [
    {"name": "Alex Khoury", "pin": "1234", "role": "tester"},
    # Add more like:
    # {"name": "John Smith", "pin": "5678", "role": "tester"},
]

# === POWER AUTOMATE WEBHOOK URL ===
# Paste your Power Automate "When HTTP request received" URL here.
# See POWER_AUTOMATE_SETUP.md for instructions.
WEBHOOK_URL = ""


def clean_columns(df, prefix):
    new_cols = {}
    for c in df.columns:
        nc = c
        if c.startswith(prefix):
            nc = c.replace(prefix, '').strip().lstrip('>').strip()
        if c.startswith('Locations >'):
            nc = 'Location'
        new_cols[c] = nc
    return df.rename(columns=new_cols)


def load_test_items(path):
    """Load TestItems sheet from TestPlan_Master.xlsm"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Warning: openpyxl not installed. Run: pip install openpyxl")
        return []

    wb = load_workbook(path, data_only=True)
    if 'TestItems' not in wb.sheetnames:
        print(f"Warning: 'TestItems' sheet not found in {path}")
        return []

    ws = wb['TestItems']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]

    items = []
    for r in range(2, ws.max_row+1):
        row = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row[h] = v
        if row.get('TestID'):
            items.append({
                'TestID': row.get('TestID', ''),
                'Phase': row.get('Phase', ''),
                'Location': row.get('Location', ''),
                'Subsystem': row.get('Subsystem', ''),
                'Activity': row.get('Activity', ''),
                'TestCategory': row.get('TestCategory', ''),
                'TestCaseCode': row.get('TestCaseCode', ''),
                'TestName': row.get('TestName', ''),
                'TestProcedure': row.get('TestProcedure', ''),
                'Status': row.get('Status', ''),
                'PlannedDate': str(row.get('PlannedDate', '')) if row.get('PlannedDate') else '',
            })
    return items


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    os.chdir(here)

    try:
        import pandas as pd
    except ImportError:
        print("ERROR: pandas not installed. Run: pip install pandas openpyxl")
        sys.exit(1)

    for f in [CSV_LINE_ITEM, CSV_ACTION_PLAN, CSV_PUNCH_LIST]:
        if not os.path.exists(f):
            print(f"ERROR: Cannot find {f}")
            sys.exit(1)

    print("Updating HITACHI Rail T&C Portal data...")
    print()

    print(f"Reading {CSV_ACTION_PLAN}...")
    ap = pd.read_csv(CSV_ACTION_PLAN)
    print(f"Reading {CSV_LINE_ITEM}...")
    li = pd.read_csv(CSV_LINE_ITEM)
    print(f"Reading {CSV_PUNCH_LIST}...")
    pl = pd.read_csv(CSV_PUNCH_LIST)

    ap = clean_columns(ap, 'Action Plan >')
    ap.columns = [c.replace('Commissioning:', '').strip() for c in ap.columns]
    li = clean_columns(li, 'Action Plan Line Item >')
    li.columns = [c.replace('Action Plan >', 'Plan').strip() for c in li.columns]
    pl = clean_columns(pl, 'Punch Item >')

    def remap_li_status(s):
        if pd.isna(s): return s
        s = str(s).strip()
        if s == 'Delayed': return 'Failed'
        if s == 'Closed': return 'Passed'
        return s

    def remap_ap_status(s):
        if pd.isna(s): return s
        s = str(s).strip()
        if s == 'Pending Test Report Acceptance': return 'Closed'
        return s

    li['Status'] = li['Status'].apply(remap_li_status)
    ap['Status'] = ap['Status'].apply(remap_ap_status)
    ap = ap[ap['Name'].notna()]
    ap = ap[ap['Phase'] != 'Grand Totals']

    ap_records = json.loads(ap.to_json(orient='records'))
    li_records = json.loads(li.to_json(orient='records'))
    pl_records = json.loads(pl.to_json(orient='records'))

    test_items = []
    if os.path.exists(XLSM_TESTPLAN):
        print(f"Reading {XLSM_TESTPLAN}...")
        test_items = load_test_items(XLSM_TESTPLAN)
    else:
        print(f"Warning: {XLSM_TESTPLAN} not found - TestItems will be empty")

    content = f"""// HITACHI Rail T&C Portal - Data
// Auto-generated. Updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

window.PORTAL_DATA = {{
  actionPlans: {json.dumps(ap_records, ensure_ascii=False)},
  lineItems: {json.dumps(li_records, ensure_ascii=False)},
  punchList: {json.dumps(pl_records, ensure_ascii=False)},
  testItems: {json.dumps(test_items, ensure_ascii=False)},
  org: {json.dumps(ORG, ensure_ascii=False)},
  fieldUsers: {json.dumps(FIELD_USERS, ensure_ascii=False)},
  config: {{
    webhookUrl: {json.dumps(WEBHOOK_URL)},
  }},
  meta: {{
    generated: "{datetime.now().isoformat()}",
    project: "BART CBTC",
    client: "Bay Area Rapid Transit",
  }}
}};
"""

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(content)

    size_kb = os.path.getsize(OUTPUT_FILE) / 1024

    print()
    print("Data updated successfully!")
    print(f"   {len(ap_records)} SAT Activities")
    print(f"   {len(li_records):,} Line Items")
    print(f"   {len(pl_records):,} Punch Items")
    print(f"   {len(test_items)} TestItems (for Field Logging)")
    print(f"   {len(FIELD_USERS)} Field Users")
    print(f"   Webhook: {'configured' if WEBHOOK_URL else 'NOT YET SET - see POWER_AUTOMATE_SETUP.md'}")
    print(f"   Output: {OUTPUT_FILE} ({size_kb:.1f} KB)")
    print()
    print("Next: Upload data.js to your GitHub repository")


if __name__ == "__main__":
    main()
